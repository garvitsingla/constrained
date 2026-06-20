import os
import random
import builtins
import io
import warnings
import logging
warnings.filterwarnings("ignore")
logging.getLogger("gymnasium").setLevel(logging.ERROR)

import torch
import numpy as np
from collections import OrderedDict
import argparse
from contextlib import contextmanager, redirect_stdout, redirect_stderr
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from torch.nn.utils.convert_parameters import parameters_to_vector
from environment import HAZARD_TYPES
import itertools
from environment import (ConstrainedGoToLocalEnv,
                         ConstrainedPickupDistEnv,
                         ConstrainedGoToObjDoorEnv,
                         ConstrainedOpenDoorEnv,
                         ConstrainedOpenDoorLocEnv,
                         ConstrainedOpenDoorsOrderEnv,
                         ConstrainedActionObjDoorEnv,
                         ConstrainedGoToOpenEnv,
                         ConstrainedFindObjS5Env)
from sampler_lang import (BabyAIMissionTaskWrapper, 
                        SentenceMissionEncoder,
                        MissionParamAdapter, 
                        ConstraintParamAdapter,
                        ConstrainedNN)
import sampler_lang
from maml_rl.policies.categorical_mlp import CategoricalMLPPolicy

# ── Helpers ────────────────────────────────────────────────────────────────────
@contextmanager
def silence():
    real_print = builtins.print
    buf = io.StringIO()
    def fp(*args, **kwargs):
        if args and isinstance(args[0], str) and "Sampling rejected" in args[0]:
            return
        real_print(*args, **kwargs)
    builtins.print = fp
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            yield
    finally:
        builtins.print = real_print

# argparser
p = argparse.ArgumentParser()
p.add_argument("--env", dest="env_name",
               choices=["ConstrainedGoToLocal","ConstrainedPickupDist","ConstrainedGoToObjDoor",
                        "ConstrainedOpenDoor","ConstrainedOpenDoorLoc","ConstrainedOpenDoorsOrder",
                        "ConstrainedActionObjDoor","ConstrainedGoToOpen","ConstrainedFindObjS5"],
               default="ConstrainedGoToLocal")
p.add_argument("--room-size", type=int, default=8)
p.add_argument("--num-dists", type=int, default=2)
p.add_argument("--max-steps", type=int, default=300)
p.add_argument("--delta-theta", type=float, default=0.3)
p.add_argument("--delta-constraint", type=float, default=0.1)
p.add_argument("--n-missions", type=int, default=10)
p.add_argument("--n-episodes", type=int, default=10)
p.add_argument("--skip-clamaml", action="store_true")
p.add_argument("--skip-unified", action="store_true")
p.add_argument("--skip-nn",      action="store_true")
p.add_argument("--skip-random",  action="store_true")
p.add_argument("--num-constraints", type=int, default=1)
args = p.parse_args()

seed = 42
random.seed(seed); np.random.seed(seed); torch.manual_seed(seed)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(seed)

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# ── Eval mission lists (eval colors/objects differ from train) ─────────────────
OBJECTS     = ['box']
COLORS      = ['red','green','blue','purple','yellow','grey']
PREP_LOCS   = ['on','at','to']
LOC_NAMES   = ['right','front']
DOOR_COLORS = ['yellow','grey']

# constraints
CONSTRAINT_TEXTS = [f"avoid {hazard}" for hazard in HAZARD_TYPES]
DOUBLE_CONSTRAINT_TEXTS = [f"avoid {h1} and avoid {h2}" for h1, h2 in itertools.combinations(HAZARD_TYPES.keys(), 2)]

# goals
LOCAL_MISSIONS    = [f"go to the {c} {o}" for c in COLORS for o in OBJECTS]
PICKUP_MISSIONS   = [f"pick up the {c} {o}" for c in COLORS for o in OBJECTS]
DOOR_MISSIONS     = [f"go to the {c} door" for c in DOOR_COLORS]
OPENDOOR_MISSIONS = [f"open the {c} door" for c in DOOR_COLORS]
OPENDOORLOC_MISSIONS  = [f"open the door {p} the {l}" for p in PREP_LOCS for l in LOC_NAMES]
OPENDOORSORDER_MISSIONS = (
    [f"open the {c} door" for c in DOOR_COLORS] +
    [f"open the {c1} door, then open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS] +
    [f"open the {c1} door after you open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
)
ACTIONOBJDOOR_MISSIONS = (
    [f"pick up the {c} {t}" for c in COLORS for t in ["box"]] +
    [f"go to the {c} {t}" for c in COLORS for t in ["box"]] +
    [f"go to the {c} door" for c in DOOR_COLORS] +
    [f"open the {c} door" for c in DOOR_COLORS]
)
FINDOBJS5_MISSIONS = [f"pick up the {t}" for t in ["box"]]


GOALS_MAP = {
    "ConstrainedGoToLocal":      LOCAL_MISSIONS,
    "ConstrainedPickupDist":     PICKUP_MISSIONS,
    "ConstrainedGoToObjDoor":    LOCAL_MISSIONS+DOOR_MISSIONS,
    "ConstrainedGoToOpen":       LOCAL_MISSIONS,
    "ConstrainedOpenDoor":       OPENDOOR_MISSIONS,
    "ConstrainedOpenDoorLoc":    OPENDOOR_MISSIONS+OPENDOORLOC_MISSIONS,
    "ConstrainedOpenDoorsOrder": OPENDOORSORDER_MISSIONS,
    "ConstrainedActionObjDoor":  ACTIONOBJDOOR_MISSIONS,
    "ConstrainedFindObjS5":      FINDOBJS5_MISSIONS
}

# ── Build environment ──────────────────────────────────────────────────────────
def build_env(env_name, room_size, num_dists, max_steps, missions,
              goals=None, constraints=None):
    dispatch = {
        "ConstrainedGoToLocal":     lambda: ConstrainedGoToLocalEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedPickupDist":    lambda: ConstrainedPickupDistEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedGoToObjDoor":   lambda: ConstrainedGoToObjDoorEnv(max_steps=max_steps, num_distractors=num_dists),
        "ConstrainedGoToOpen":      lambda: ConstrainedGoToOpenEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedOpenDoor":      lambda: ConstrainedOpenDoorEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedOpenDoorLoc":   lambda: ConstrainedOpenDoorLocEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedOpenDoorsOrder":lambda: ConstrainedOpenDoorsOrderEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedActionObjDoor": lambda: ConstrainedActionObjDoorEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedFindObjS5":     lambda: ConstrainedFindObjS5Env(room_size=5, max_steps=max_steps),
    }
    base = dispatch[env_name]()
    return BabyAIMissionTaskWrapper(base, missions=missions, goals=goals, constraints=constraints)


# ── Setup: env + shared encoder ───────────────────────────────────────────────
env_name    = args.env_name
room_size   = "env" if env_name in ["GoToObjDoor"] else args.room_size
num_dists   = "env" if env_name in ["OpenDoor","OpenDoorLoc","OpenDoorsOrder"] else args.num_dists
max_steps   = args.max_steps
delta_theta = args.delta_theta
delta_c     = args.delta_constraint
n_missions  = args.n_missions
n_episodes          = args.n_episodes

goals_list = GOALS_MAP.get(env_name)
if args.num_constraints == 1:
    constraints_list = CONSTRAINT_TEXTS
    all_missions = [f"{g} and {c}" for g in goals_list for c in CONSTRAINT_TEXTS]
else:
    constraints_list = DOUBLE_CONSTRAINT_TEXTS
    all_missions = [f"{g} and {c}" for g in goals_list for c in DOUBLE_CONSTRAINT_TEXTS]

env = build_env(env_name, room_size, num_dists, max_steps, all_missions,
                goals_list, constraints_list)

print(f"\n{'='*65}")
print(f"Evaluation: {env_name}")
print(f"Tasks: {n_missions} | Episodes per task: {n_episodes} | delta_theta: {delta_theta} | delta constraint: {delta_c}")
print(f"{'='*65}\n")

dummy_obs, _ = env.reset()
input_size   = sampler_lang.preprocess_obs(dummy_obs).shape[0]
output_size  = env.action_space.n
hidden_sizes = (64, 64)
non_linearity = torch.nn.functional.tanh

mission_encoder = SentenceMissionEncoder(
    model_name="all-MiniLM-L6-v2", 
    frozen=True, 
    normalize=True, 
    cache=True, 
    device=device
)
mission_encoder.eval()
enc_dim = mission_encoder.output_dim

def _make_policy():
    pol = CategoricalMLPPolicy(input_size=input_size, 
                               output_size=output_size,
                               hidden_sizes=hidden_sizes, 
                               nonlinearity=non_linearity).to(device)
    return pol

policy_param_shapes = [p.shape for p in _make_policy().parameters()]


# ── Load C-LAMAML ─────────────────────────────────────────────────────────────
clamaml_ready = False
if not args.skip_clamaml:
    _ckpt_path = f"lang_model/lang_{env_name}_dt{delta_theta}_dc{delta_c}_{args.num_constraints}c.pth"
    if os.path.exists(_ckpt_path):
        ckpt_c = torch.load(_ckpt_path, map_location=device)
        policy_c = _make_policy(); policy_c.load_state_dict(ckpt_c["policy"]); policy_c.eval()
        adapter_c = MissionParamAdapter(enc_dim, policy_param_shapes).to(device)
        adapter_c.load_state_dict(ckpt_c["mission_adapter"]); adapter_c.eval()
        constr_adapter = None
        if "constraint_adapter" in ckpt_c:
            constr_adapter = ConstraintParamAdapter(enc_dim, policy_param_shapes).to(device)
            constr_adapter.load_state_dict(ckpt_c["constraint_adapter"]); constr_adapter.eval()
        clamaml_ready = True
        print(f"[✓] C-LAMAML loaded from {_ckpt_path}")
    else:
        print(f"[✗] C-LAMAML checkpoint not found: {_ckpt_path}  (skipping)")


# ── LoadLA-MAML ───────────────────────────────────────────────────────
unified_ready = False
if not args.skip_unified:
    _ckpt_path = f"unified_model/lang_{env_name}_{delta_theta}_{args.num_constraints}c.pth"
    if os.path.exists(_ckpt_path):
        ckpt_u = torch.load(_ckpt_path, map_location=device)
        policy_u = _make_policy(); policy_u.load_state_dict(ckpt_u["policy"]); policy_u.eval()
        adapter_u = MissionParamAdapter(enc_dim, policy_param_shapes).to(device)
        adapter_u.load_state_dict(ckpt_u["mission_adapter"]); adapter_u.eval()
        unified_ready = True
        print(f"[✓] LA-MAML loaded from {_ckpt_path}")
    else:
        print(f"[✗] LA-MAML checkpoint not found: {_ckpt_path}  (skipping)")


# ── Load NN C-LAMAML ───────────────────────────────────────────────────────
nn_ready = False
if not args.skip_nn:
    _ckpt_path = f"nn_model/lang_{env_name}_nn_{args.num_constraints}c.pth"
    if os.path.exists(_ckpt_path):
        ckpt_h = torch.load(_ckpt_path, map_location=device)
        policy_n = _make_policy(); policy_n.load_state_dict(ckpt_h["policy"]); policy_n.eval()
        nn_net = None
        if ckpt_h.get("nn") is not None:
            nn_net = ConstrainedNN(enc_dim, policy_param_shapes).to(device)
            nn_net.load_state_dict(ckpt_h["nn"]); nn_net.eval()
        nn_ready = True
        print(f"[✓] NN C-LAMAML loaded from {_ckpt_path}")
    else:
        print(f"[✗] NN C-LAMAML checkpoint not found: {_ckpt_path}  (skipping)")

print()

# ── Adapted params helpers ─────────────────────────────────────────────────────
def _params_clamaml(mission):
    """θ' = θ + Δθ_goal + Δθ_constraint"""
    parts = mission.split(" and avoid ", 1)
    goal_str, constr_str = (parts[0], "avoid " + parts[1]) if len(parts) == 2 else (mission, None)

    with torch.no_grad():
        g_emb   = mission_encoder(goal_str).to(device)
        deltas_g = adapter_c(g_emb)
        deltas_c  = [torch.zeros_like(d) for d in deltas_g]
        if constr_str and constr_adapter:
            c_emb    = mission_encoder(constr_str).to(device)
            deltas_c = constr_adapter(c_emb)

        names  = list(dict(policy_c.named_parameters()).keys())
        params = list(policy_c.parameters())
        return OrderedDict(
            (n, p + dg.squeeze(0) * delta_theta + dc.squeeze(0) * delta_c)
            for n, p, dg, dc in zip(names, params, deltas_g, deltas_c)
        )


def _params_unified(mission):
    """θ' = θ + Δθ_unified  (full combined string)"""
    combined = f"{mission[0]} and {mission[1]}" if isinstance(mission, tuple) else mission
    with torch.no_grad():
        emb    = mission_encoder(combined).to(device)
        deltas = adapter_u(emb)
        names  = list(dict(policy_u.named_parameters()).keys())
        params = list(policy_u.parameters())
        return OrderedDict(
            (n, p + d.squeeze(0) * delta_theta)
            for n, p, d in zip(names, params, deltas)
        )


def _params_nn(mission):
    """θ' = NN(θ, goal_emb, constr_emb)"""
    parts = mission.split(" and avoid ", 1)
    goal_str, constr_str = (parts[0], "avoid " + parts[1]) if len(parts) == 2 else (mission, None)

    with torch.no_grad():
        g_emb = mission_encoder(goal_str).to(device)
        c_emb = mission_encoder(constr_str).to(device) if constr_str else torch.zeros_like(g_emb)

        if nn_net is not None:
            theta_flat    = parameters_to_vector(list(policy_n.parameters()))
            combined_inp  = torch.cat([theta_flat.unsqueeze(0), g_emb, c_emb], dim=-1)
            theta_tensors = nn_net(combined_inp)
            names = list(dict(policy_n.named_parameters()).keys())
            return OrderedDict((n, t.squeeze(0)) for n, t in zip(names, theta_tensors))
        else:
            return OrderedDict(policy_n.named_parameters())


# ── Single-episode rollout ─────────────────────────────────────────────────────
def rollout(policy, params, preproc=sampler_lang.preprocess_obs, seed=None):
    with silence():
        obs, _ = env.reset(seed=seed)
    done, steps, success, viols = False, 0, False, 0
    env_max = getattr(env.unwrapped, 'max_steps', max_steps)
    while not done and steps < env_max:
        obs_t = torch.from_numpy(preproc(obs)[None]).float().to(device)
        with torch.no_grad():
            action = policy(obs_t, params=params).sample().item()
        obs, reward, terminated, truncated, info = env.step(action)
        done = terminated or truncated
        steps += 1
        viols += int(info.get('cost', 0) > 0)
        if terminated:
            success = True
    return steps, success, viols


def rollout_random(seed=None):
    with silence():
        obs, _ = env.reset(seed=seed)
    done, steps, success, viols = False, 0, False, 0
    env_max = getattr(env.unwrapped, 'max_steps', max_steps)
    while not done and steps < env_max:
        obs, reward, terminated, truncated, info = env.step(env.action_space.sample())
        done = terminated or truncated
        steps += 1
        viols += int(info.get('cost', 0) > 0)
        if terminated:
            success = True
    return steps, success, viols


# ── Select test tasks ──────────────────────────────────────────────────────────
test_tasks = random.sample(all_missions, min(n_missions, len(all_missions)))

# ── Run evaluation ─────────────────────────────────────────────────────────────
METHODS = []
if clamaml_ready:  METHODS.append(("C-LAMAML", policy_c, _params_clamaml))
if unified_ready:  METHODS.append(("LAMAML", policy_u, _params_unified))
if nn_ready:       METHODS.append(("NN_C_LAMAML", policy_n, _params_nn))
METHODS.append(("Random", None, None))

# per-method stats: list of (sr, avg_steps, avg_viols) per task
per_task = {m[0]: [] for m in METHODS}

COL_W = 14
header = f"{'Mission':<45}" + "".join(f"{'  SR%/Steps/Viols':>{COL_W*3}}" for m,_,__ in METHODS)
print(f"\n{'Mission':<45} " +
      " | ".join(f"{'SR%':>5} {'Steps':>6} {'Viols':>5}" for m,_,__ in METHODS))
print(f"  ({'  |  '.join(m[0] for m in METHODS)})")
print("-" * (45 + 22 * len(METHODS)))

for mission in test_tasks:
    env.reset_task(mission)
    row  = f"{mission[:44]:<45}"

    ep_seeds = [random.randint(0, 1000000) for _ in range(n_episodes)]

    for (mname, policy, get_params) in METHODS:
        ep_steps, ep_succ, ep_viols = [], [], []
        if mname == "Random":
            for ep in range(n_episodes):
                env.reset_task(mission)
                s, ok, v = rollout_random(seed=ep_seeds[ep])
                ep_steps.append(s); ep_succ.append(ok); ep_viols.append(v)
        else:
            params = get_params(mission)
            for ep in range(n_episodes):
                env.reset_task(mission)
                s, ok, v = rollout(policy, params, seed=ep_seeds[ep])
                ep_steps.append(s); ep_succ.append(ok); ep_viols.append(v)
        
        sr    = np.mean(ep_succ)
        avgst = np.mean(ep_steps)
        avgv  = np.mean(ep_viols)
        per_task[mname].append((sr, avgst, avgv))
        row += f" | {sr*100:>4.0f}% {avgst:>6.1f} {avgv:>5.1f}"

    print(row)

# ── Summary ────────────────────────────────────────────────────────────────────
print("\n" + "="*65)
print(f"{'Method':<18} | {'Avg SR%':>7} | {'Avg Steps':>9} | {'Avg Viols':>9}")
print("-"*65)
for (mname, _, __) in METHODS:
    data = per_task[mname]
    srs   = [x[0] for x in data]
    steps = [x[1] for x in data]
    viols = [x[2] for x in data]
    print(f"{mname:<18} | {np.mean(srs)*100:>6.1f}% | {np.mean(steps):>9.1f} | {np.mean(viols):>9.2f}")
print("="*65)

# ── Excel logging ──────────────────────────────────────────────────────────────
xlsx_path = "evaluation_results.xlsx"
if os.path.exists(xlsx_path):
    wb = load_workbook(xlsx_path)
else:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

# Summary sheet (append one row per run)
summary_name = "Summary"
if summary_name not in wb.sheetnames:
    ws_sum = wb.create_sheet(summary_name)
    # Row 1: Merged Method Names
    ws_sum.append(["Env", "Room", "Dists", "Steps", 
                   "Constrained-LAMAML", "", "", 
                   "LAMAML", "", "", 
                   "NN_C_LAMAML", "", "", 
                   "Random", "", ""])
    
    # Row 2: Specific Metrics
    ws_sum.append(["", "", "", "", 
                   "SR%", "Steps", "Viols", 
                   "SR%", "Steps", "Viols", 
                   "SR%", "Steps", "Viols", 
                   "SR%", "Steps", "Viols"])

    # Merges
    ws_sum.merge_cells('E1:G1') # Constrained-LAMAML
    ws_sum.merge_cells('H1:J1') # LAMAML
    ws_sum.merge_cells('K1:M1') # NN_C_LAMAML
    ws_sum.merge_cells('N1:P1') # Random
    
    ws_sum.merge_cells('A1:A2') # Env
    ws_sum.merge_cells('B1:B2') # Room
    ws_sum.merge_cells('C1:C2') # Dists
    ws_sum.merge_cells('D1:D2') # Steps

    # Styling
    center = Alignment(horizontal='center', vertical='center')
    for row in ws_sum.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = center
            cell.font = Font(bold=True)
else:
    ws_sum = wb[summary_name]

def _agg(mname):
    data = per_task.get(mname, [])
    if not data:
        return "", "", ""
    return (round(np.mean([x[0] for x in data])*100, 1),
            round(np.mean([x[1] for x in data]), 1),
            round(np.mean([x[2] for x in data]), 2))

ws_sum.append([f"{env_name} ({args.num_constraints}c)", room_size, num_dists, max_steps,
               *_agg("C-LAMAML"), *_agg("LAMAML"), *_agg("NN_C_LAMAML"), *_agg("Random")])

# Per-mission sheet (overwritten per run)
sheet_name = (f"{env_name}_{args.num_constraints}c_Missions")[:31]
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws_m = wb.create_sheet(sheet_name)
header_row = ["Mission"]
for (mname, _, __) in METHODS:
    header_row += [f"{mname} SR%", f"{mname} Steps", f"{mname} Viols"]
ws_m.append(header_row)

for i, mission in enumerate(test_tasks):
    row = [mission]
    for (mname, _, __) in METHODS:
        sr, st, vl = per_task[mname][i]
        row += [round(sr*100,1), round(st,1), round(vl,2)]
    ws_m.append(row)

# OVERALL row
overall_row = ["OVERALL"]
for (mname, _, __) in METHODS:
    sr, st, vl = _agg(mname)
    overall_row += [sr, st, vl]
ws_m.append(overall_row)

# Make the last row bold
bold_font = Font(bold=True)
for cell in ws_m[ws_m.max_row]:
    cell.font = bold_font

wb.save(xlsx_path)
print(f"\nResults saved → {xlsx_path}  (sheets: '{summary_name}', '{sheet_name}')")
