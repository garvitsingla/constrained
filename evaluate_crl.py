import os
import random
import builtins
import io
import warnings
import logging
import itertools

import torch
import numpy as np
from collections import OrderedDict
from contextlib import contextmanager, redirect_stdout, redirect_stderr
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

warnings.filterwarnings("ignore")
logging.getLogger("gymnasium").setLevel(logging.ERROR)

from environment import (
    HAZARD_TYPES,
    ConstrainedGoToLocalEnv, ConstrainedPickupDistEnv,
    ConstrainedGoToObjDoorEnv, ConstrainedOpenDoorEnv,
    ConstrainedOpenDoorLocEnv, ConstrainedOpenDoorsOrderEnv,
    ConstrainedActionObjDoorEnv, ConstrainedGoToOpenEnv,
    ConstrainedFindObjS5Env,
)
from sampler_lang import (
    BabyAIMissionTaskWrapper, SentenceMissionEncoder,
    MissionParamAdapter, ConstraintParamAdapter,
    preprocess_obs,
)
from maml_rl.policies.categorical_mlp import CategoricalMLPPolicy

import argparse

# ─────────────────────────────────────────────────────────────────────────────
# Argparser
# ─────────────────────────────────────────────────────────────────────────────
p = argparse.ArgumentParser()
p.add_argument("--env", dest="env_name",
               choices=["ConstrainedGoToLocal","ConstrainedPickupDist","ConstrainedGoToObjDoor",
                        "ConstrainedOpenDoor","ConstrainedOpenDoorLoc","ConstrainedOpenDoorsOrder",
                        "ConstrainedActionObjDoor","ConstrainedGoToOpen","ConstrainedFindObjS5"],
               default="ConstrainedGoToLocal")
p.add_argument("--room-size", type=int, default=8)
p.add_argument("--num-dists", type=int, default=2)
p.add_argument("--max-steps", type=int, default=300)
p.add_argument("--delta-theta", type=float, default=0.3)
p.add_argument("--delta-constraint", type=float, default=0.1)
p.add_argument("--n-missions", type=int, default=10)
p.add_argument("--n-episodes", type=int, default=10)
p.add_argument("--num-constraints", type=int, default=1)
p.add_argument("--skip-random", action="store_true")
args = p.parse_args()

# ─────────────────────────────────────────────────────────────────────────────
# Seed
# ─────────────────────────────────────────────────────────────────────────────
seed = 42
random.seed(seed)
np.random.seed(seed)
torch.manual_seed(seed)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(seed)

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


# ─────────────────────────────────────────────────────────────────────────────
# Silence helper
# ─────────────────────────────────────────────────────────────────────────────
@contextmanager
def silence():
    real_print = builtins.print
    buf = io.StringIO()
    def fp(*args, **kwargs):
        if args and isinstance(args[0], str) and "Sampling rejected" in args[0]:
            return
        real_print(*args, **kwargs)
    builtins.print = fp
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            yield
    finally:
        builtins.print = real_print


# ─────────────────────────────────────────────────────────────────────────────
# Mission definitions
# ─────────────────────────────────────────────────────────────────────────────
OBJECTS     = ['box']
COLORS      = ['red','green','blue','purple','yellow','grey']
PREP_LOCS   = ['on','at','to']
LOC_NAMES   = ['right','front']
DOOR_COLORS = ['yellow','grey']

CONSTRAINT_TEXTS = [f"avoid {h}" for h in HAZARD_TYPES]
DOUBLE_CONSTRAINT_TEXTS = [
    f"avoid {h1} and avoid {h2}"
    for h1, h2 in itertools.combinations(HAZARD_TYPES.keys(), 2)
]

LOCAL_MISSIONS    = [f"go to the {c} {o}" for c in COLORS for o in OBJECTS]
PICKUP_MISSIONS   = [f"pick up the {c} {o}" for c in COLORS for o in OBJECTS]
DOOR_MISSIONS     = [f"go to the {c} door" for c in DOOR_COLORS]
OPENDOOR_MISSIONS = [f"open the {c} door" for c in DOOR_COLORS]
OPENDOORLOC_MISSIONS = [f"open the door {p} the {l}" for p in PREP_LOCS for l in LOC_NAMES]
OPENDOORSORDER_MISSIONS = (
    [f"open the {c} door" for c in DOOR_COLORS] +
    [f"open the {c1} door, then open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS] +
    [f"open the {c1} door after you open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
)
ACTIONOBJDOOR_MISSIONS = (
    [f"pick up the {c} {t}" for c in COLORS for t in ["box"]] +
    [f"go to the {c} {t}" for c in COLORS for t in ["box"]] +
    [f"go to the {c} door" for c in DOOR_COLORS] +
    [f"open the {c} door" for c in DOOR_COLORS]
)
FINDOBJS5_MISSIONS = [f"pick up the {t}" for t in ["box"]]

GOALS_MAP = {
    "ConstrainedGoToLocal":      LOCAL_MISSIONS,
    "ConstrainedPickupDist":     PICKUP_MISSIONS,
    "ConstrainedGoToObjDoor":    LOCAL_MISSIONS + DOOR_MISSIONS,
    "ConstrainedGoToOpen":       LOCAL_MISSIONS,
    "ConstrainedOpenDoor":       OPENDOOR_MISSIONS,
    "ConstrainedOpenDoorLoc":    OPENDOOR_MISSIONS + OPENDOORLOC_MISSIONS,
    "ConstrainedOpenDoorsOrder": OPENDOORSORDER_MISSIONS,
    "ConstrainedActionObjDoor":  ACTIONOBJDOOR_MISSIONS,
    "ConstrainedFindObjS5":      FINDOBJS5_MISSIONS,
}


# ─────────────────────────────────────────────────────────────────────────────
# Environment builder
# ─────────────────────────────────────────────────────────────────────────────
def build_env(env_name, room_size, num_dists, max_steps, missions, goals, constraints):
    room_size = room_size if room_size != "env" else args.room_size
    num_dists = num_dists if num_dists != "env" else args.num_dists

    dispatch = {
        "ConstrainedGoToLocal":      lambda: ConstrainedGoToLocalEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedPickupDist":     lambda: ConstrainedPickupDistEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedGoToObjDoor":    lambda: ConstrainedGoToObjDoorEnv(max_steps=max_steps, num_distractors=num_dists),
        "ConstrainedGoToOpen":       lambda: ConstrainedGoToOpenEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedOpenDoor":       lambda: ConstrainedOpenDoorEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedOpenDoorLoc":    lambda: ConstrainedOpenDoorLocEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedOpenDoorsOrder": lambda: ConstrainedOpenDoorsOrderEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedActionObjDoor":  lambda: ConstrainedActionObjDoorEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedFindObjS5":      lambda: ConstrainedFindObjS5Env(room_size=5, max_steps=max_steps),
    }
    base = dispatch[env_name]()
    return BabyAIMissionTaskWrapper(base, missions=missions, goals=goals, constraints=constraints)


# ─────────────────────────────────────────────────────────────────────────────
# Evaluation configs per environment
# ─────────────────────────────────────────────────────────────────────────────
def get_configs(env_name):
    if env_name in ["ConstrainedGoToLocal", "ConstrainedPickupDist"]:
        return [(7, 3), (7, 5), (8, 2), (8, 4), (9, 3), (9, 5)]
    elif env_name == "ConstrainedGoToObjDoor":
        return [("env", 1), ("env", 2), ("env", 3), ("env", 4), ("env", 5)]
    elif env_name == "ConstrainedActionObjDoor":
        return [("env", "env")]
    elif env_name in ["ConstrainedGoToOpen", "ConstrainedFindObjS5"]:
        return [(5, 2), (5, 3), (6, 2), (6, 4)]
    else:
        return [(6, "env"), (7, "env"), (8, "env"), (9, "env"), (10, "env")]


# ─────────────────────────────────────────────────────────────────────────────
# Single-episode rollout
# ─────────────────────────────────────────────────────────────────────────────
def evaluate_policy(env, policy, params=None, seed=None):
    """Run one episode. Returns (steps, success, violations)."""
    with silence():
        obs, _ = env.reset(seed=seed)
    done, steps, success, viols = False, 0, False, 0
    env_max = getattr(env.unwrapped, 'max_steps', args.max_steps)

    while not done and steps < env_max:
        obs_t = torch.from_numpy(preprocess_obs(obs)[None]).float().to(device)
        with torch.no_grad():
            if params is not None:
                action = policy(obs_t, params=params).sample().item()
            else:
                action = policy(obs_t).sample().item()
        obs, reward, terminated, truncated, info = env.step(action)
        done = terminated or truncated
        steps += 1
        viols += int(info.get('cost', 0) > 0)
        if terminated:
            success = True

    return steps, success, viols


def evaluate_random(env, seed=None):
    """Run one episode with a random action policy."""
    with silence():
        obs, _ = env.reset(seed=seed)
    done, steps, success, viols = False, 0, False, 0
    env_max = getattr(env.unwrapped, 'max_steps', args.max_steps)

    while not done and steps < env_max:
        obs, reward, terminated, truncated, info = env.step(env.action_space.sample())
        done = terminated or truncated
        steps += 1
        viols += int(info.get('cost', 0) > 0)
        if terminated:
            success = True

    return steps, success, viols


# ─────────────────────────────────────────────────────────────────────────────
# Adapted params for C-LAMAML
# ─────────────────────────────────────────────────────────────────────────────
def get_clamaml_params(mission, policy, encoder, m_adapter, c_adapter, delta_theta, delta_c):
    """θ' = θ + Δθ_goal * delta_theta + Δθ_constraint * delta_c"""
    parts = mission.split(" and avoid ", 1)
    goal_str = parts[0]
    constr_str = ("avoid " + parts[1]) if len(parts) == 2 else None

    with torch.no_grad():
        g_emb    = encoder(goal_str).to(device)
        deltas_g = m_adapter(g_emb)
        deltas_c = [torch.zeros_like(d) for d in deltas_g]
        if constr_str and c_adapter:
            c_emb    = encoder(constr_str).to(device)
            deltas_c = c_adapter(c_emb)

        names  = list(dict(policy.named_parameters()).keys())
        params = list(policy.parameters())
        return OrderedDict(
            (n, p + dg.squeeze(0) * delta_theta + dc.squeeze(0) * delta_c)
            for n, p, dg, dc in zip(names, params, deltas_g, deltas_c)
        )


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
env_name    = args.env_name
max_steps   = args.max_steps
delta_theta = args.delta_theta
delta_c     = args.delta_constraint
n_missions  = args.n_missions
n_episodes  = args.n_episodes
nc          = args.num_constraints

# Build mission lists
goals_list = GOALS_MAP[env_name]
constraints_list = CONSTRAINT_TEXTS if nc == 1 else DOUBLE_CONSTRAINT_TEXTS
all_missions = [f"{g} and {c}" for g in goals_list for c in constraints_list]

# Create dummy env for observation/action shapes
dummy_env = build_env(env_name, args.room_size, args.num_dists, max_steps,
                      all_missions, goals_list, constraints_list)
dummy_obs, _ = dummy_env.reset()
input_size   = preprocess_obs(dummy_obs).shape[0]
output_size  = dummy_env.action_space.n
hidden_sizes = (64, 64)
nonlinearity = torch.nn.functional.tanh

def make_policy():
    return CategoricalMLPPolicy(
        input_size=input_size, output_size=output_size,
        hidden_sizes=hidden_sizes, nonlinearity=nonlinearity,
    ).to(device)

policy_param_shapes = [p.shape for p in make_policy().parameters()]


# ─────────────────────────────────────────────────────────────────────────────
# 1. Load C-LAMAML
# ─────────────────────────────────────────────────────────────────────────────
clamaml_ckpt = f"lang_model/lang_{env_name}_dt{delta_theta}_dc{delta_c}_{nc}c.pth"
if os.path.exists(clamaml_ckpt):
    ckpt = torch.load(clamaml_ckpt, map_location=device)
    policy_clamaml = make_policy()
    policy_clamaml.load_state_dict(ckpt["policy"])
    policy_clamaml.eval()

    encoder = SentenceMissionEncoder(
        model_name="all-MiniLM-L6-v2", frozen=True,
        normalize=True, cache=True, device=device,
    )
    encoder.eval()
    enc_dim = encoder.output_dim

    m_adapter = MissionParamAdapter(enc_dim, policy_param_shapes).to(device)
    m_adapter.load_state_dict(ckpt["mission_adapter"])
    m_adapter.eval()

    c_adapter = None
    if "constraint_adapter" in ckpt:
        c_adapter = ConstraintParamAdapter(enc_dim, policy_param_shapes).to(device)
        c_adapter.load_state_dict(ckpt["constraint_adapter"])
        c_adapter.eval()

    print(f"[✓] C-LAMAML loaded from {clamaml_ckpt}")
else:
    raise FileNotFoundError(f"C-LAMAML checkpoint not found: {clamaml_ckpt}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. Load Constrained RL
# ─────────────────────────────────────────────────────────────────────────────
crl_ckpt = f"crl_model/crl_{env_name}_{nc}c.pth"
if os.path.exists(crl_ckpt):
    ckpt_crl = torch.load(crl_ckpt, map_location=device)
    policy_crl = make_policy()
    policy_crl.load_state_dict(ckpt_crl["policy"])
    policy_crl.eval()
    crl_ready = True
    print(f"[✓] Constrained RL loaded from {crl_ckpt}")
else:
    crl_ready = False
    print(f"[✗] Constrained RL checkpoint not found: {crl_ckpt}  (skipping)")


# ─────────────────────────────────────────────────────────────────────────────
# 3. Random baseline (no loading needed)
# ─────────────────────────────────────────────────────────────────────────────
random_ready = not args.skip_random


# ─────────────────────────────────────────────────────────────────────────────
# Evaluation loop
# ─────────────────────────────────────────────────────────────────────────────
configs    = get_configs(env_name)
test_tasks = random.sample(all_missions, min(n_missions, len(all_missions)))

# Accumulators across all configs
results_clamaml = {'steps': [], 'successes': [], 'viols': []}
results_crl     = {'steps': [], 'successes': [], 'viols': []}
results_random  = {'steps': [], 'successes': [], 'viols': []}

excel_rows = []

print(f"\n{'='*65}")
print(f"Evaluation: {env_name}")
print(f"Tasks: {n_missions} | Episodes/task: {n_episodes} | "
      f"delta_theta: {delta_theta} | delta_c: {delta_c}")
print(f"Methods: C-LAMAML | {'Constrained RL | ' if crl_ready else ''}{'Random' if random_ready else ''}")
print(f"{'='*65}\n")


for c_room, c_dists in configs:
    print(f"Config: Room={c_room}, Dists={c_dists}")

    env = build_env(env_name, c_room, c_dists, max_steps,
                    all_missions, goals_list, constraints_list)

    cfg_clamaml = {'steps': [], 'successes': [], 'viols': []}
    cfg_crl     = {'steps': [], 'successes': [], 'viols': []}
    cfg_random  = {'steps': [], 'successes': [], 'viols': []}

    for mission in test_tasks:
        ep_seeds = [random.randint(0, 1_000_000) for _ in range(n_episodes)]

        # ── C-LAMAML ────────────────────────────────────────────────────
        theta_prime = get_clamaml_params(
            mission, policy_clamaml, encoder, m_adapter, c_adapter,
            delta_theta, delta_c
        )
        for ep in range(n_episodes):
            env.reset_task(mission)
            s, ok, v = evaluate_policy(env, policy_clamaml, params=theta_prime, seed=ep_seeds[ep])
            cfg_clamaml['steps'].append(s)
            cfg_clamaml['successes'].append(ok)
            cfg_clamaml['viols'].append(v)

        # ── Constrained RL ──────────────────────────────────────────────
        if crl_ready:
            for ep in range(n_episodes):
                env.reset_task(mission)
                s, ok, v = evaluate_policy(env, policy_crl, params=None, seed=ep_seeds[ep])
                cfg_crl['steps'].append(s)
                cfg_crl['successes'].append(ok)
                cfg_crl['viols'].append(v)

        # ── Random ──────────────────────────────────────────────────────
        if random_ready:
            for ep in range(n_episodes):
                env.reset_task(mission)
                s, ok, v = evaluate_random(env, seed=ep_seeds[ep])
                cfg_random['steps'].append(s)
                cfg_random['successes'].append(ok)
                cfg_random['viols'].append(v)

    # Accumulate into global results
    results_clamaml['steps'].extend(cfg_clamaml['steps'])
    results_clamaml['successes'].extend(cfg_clamaml['successes'])
    results_clamaml['viols'].extend(cfg_clamaml['viols'])

    if crl_ready:
        results_crl['steps'].extend(cfg_crl['steps'])
        results_crl['successes'].extend(cfg_crl['successes'])
        results_crl['viols'].extend(cfg_crl['viols'])

    if random_ready:
        results_random['steps'].extend(cfg_random['steps'])
        results_random['successes'].extend(cfg_random['successes'])
        results_random['viols'].extend(cfg_random['viols'])

    # Build Excel row for this config
    def fmt(bucket):
        ms = np.mean(bucket['steps']) if bucket['steps'] else 0.0
        ss = np.std(bucket['steps'])  if bucket['steps'] else 0.0
        sr = round(np.mean(bucket['successes']), 2) if bucket['successes'] else 0.0
        mv = np.mean(bucket['viols']) if bucket['viols'] else 0.0
        sv = np.std(bucket['viols'])  if bucket['viols'] else 0.0
        return [f"{ms:.2f} ± {ss:.2f}", sr, f"{mv:.2f} ± {sv:.2f}"]

    row = [c_room, c_dists, max_steps, delta_theta]
    row += fmt(cfg_clamaml)
    if crl_ready:
        row += fmt(cfg_crl)
    if random_ready:
        row += fmt(cfg_random)
    excel_rows.append(row)

    # Print a quick summary for this config
    sr_c = np.mean(cfg_clamaml['successes'])*100 if cfg_clamaml['successes'] else 0
    sr_r = np.mean(cfg_crl['successes'])*100 if crl_ready and cfg_crl['successes'] else 0
    print(f"  C-LAMAML SR: {sr_c:.1f}%  |  CRL SR: {sr_r:.1f}%")


# ─────────────────────────────────────────────────────────────────────────────
# AVERAGE row
# ─────────────────────────────────────────────────────────────────────────────
def fmt_global(bucket):
    ms = np.mean(bucket['steps']) if bucket['steps'] else 0.0
    ss = np.std(bucket['steps'])  if bucket['steps'] else 0.0
    sr = round(np.mean(bucket['successes']), 2) if bucket['successes'] else 0.0
    mv = np.mean(bucket['viols']) if bucket['viols'] else 0.0
    sv = np.std(bucket['viols'])  if bucket['viols'] else 0.0
    return [f"{ms:.2f} ± {ss:.2f}", sr, f"{mv:.2f} ± {sv:.2f}"]

avg_row = ["AVERAGE", "", "", ""]
avg_row += fmt_global(results_clamaml)
if crl_ready:
    avg_row += fmt_global(results_crl)
if random_ready:
    avg_row += fmt_global(results_random)
excel_rows.append(avg_row)


# ─────────────────────────────────────────────────────────────────────────────
# Console summary
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("FINAL AGGREGATE RESULTS")
print(f"{'='*65}")
print(f"C-LAMAML       :  SR={np.mean(results_clamaml['successes'])*100:.2f}%  "
      f"Steps={np.mean(results_clamaml['steps']):.2f} ± {np.std(results_clamaml['steps']):.2f}  "
      f"Viols={np.mean(results_clamaml['viols']):.2f} ± {np.std(results_clamaml['viols']):.2f}")
if crl_ready:
    print(f"Constrained RL :  SR={np.mean(results_crl['successes'])*100:.2f}%  "
          f"Steps={np.mean(results_crl['steps']):.2f} ± {np.std(results_crl['steps']):.2f}  "
          f"Viols={np.mean(results_crl['viols']):.2f} ± {np.std(results_crl['viols']):.2f}")
if random_ready:
    print(f"Random         :  SR={np.mean(results_random['successes'])*100:.2f}%  "
          f"Steps={np.mean(results_random['steps']):.2f} ± {np.std(results_random['steps']):.2f}  "
          f"Viols={np.mean(results_random['viols']):.2f} ± {np.std(results_random['viols']):.2f}")


# ─────────────────────────────────────────────────────────────────────────────
# Save to Excel
# ─────────────────────────────────────────────────────────────────────────────
xlsx_path = "crl_results.xlsx"
if os.path.exists(xlsx_path):
    wb = load_workbook(xlsx_path)
else:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

sheet_name = f"{env_name}_{nc}c"[:31]
is_new_sheet = False
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.append([])
    ws.append([f"--- NEW RUN: C-LAMAML (dt={delta_theta}, dc={delta_c}) vs CRL vs Random ---"])
    ws[ws.max_row][0].font = Font(bold=True)
else:
    ws = wb.create_sheet(sheet_name)
    is_new_sheet = True

# Header
header = ["Room Size", "Num Distractor", "Max Steps", "Delta Theta",
          "Avg Steps C-LAMAML", "Success Prob C-LAMAML", "Avg Viols C-LAMAML"]
if crl_ready:
    header += ["Avg Steps CRL", "Success Prob CRL", "Avg Viols CRL"]
if random_ready:
    header += ["Avg Steps Random", "Success Prob Random", "Avg Viols Random"]
ws.append(header)

# Bold the header row
for cell in ws[ws.max_row]:
    cell.font = Font(bold=True)

for row in excel_rows:
    ws.append(row)

# Bold the AVERAGE row
for cell in ws[ws.max_row]:
    cell.font = Font(bold=True)

wb.save(xlsx_path)
print(f"\nResults saved → {xlsx_path}  (sheet: '{sheet_name}')")
