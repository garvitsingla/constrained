"""
Constraint RL scratch baseline for C-LAMAML.

Purpose
-------
For each unseen evaluation mission, compare:
  1. C-LAMAML zero/few-step adapted policy loaded from lang_model/...
  2. A scratch-trained non-meta Constraint RL policy trained only on that same mission

The scratch baseline answers:
  "How many updates / seconds / environment episodes does ordinary constrained RL need
   to match the already-trained C-LAMAML performance on the same unseen task?"

This is intentionally NOT meta-learning:
  - no mission adapter
  - no constraint adapter
  - no theta prime from language
  - each unseen mission starts from a randomly initialized policy
  - training is done directly on the single target mission using a penalized objective

Run example
-----------
python constraint_rl_baseline.py \
  --env ConstrainedGoToLocal \
  --room-size 8 \
  --num-dists 2 \
  --max-steps 300 \
  --delta-theta 0.3 \
  --delta-constraint 0.1 \
  --num-constraints 1 \
  --n-missions 10 \
  --eval-episodes 10 \
  --scratch-iters 300 \
  --scratch-batch-size 20
"""

import os
import time
import json
import random
import argparse
import warnings
import logging
from collections import OrderedDict
from contextlib import contextmanager, redirect_stdout, redirect_stderr
import builtins
import io
import itertools

import numpy as np
import torch
import torch.nn.functional as F
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

warnings.filterwarnings("ignore")
logging.getLogger("gymnasium").setLevel(logging.ERROR)
os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")

from environment import (
    HAZARD_TYPES,
    ConstrainedGoToLocalEnv,
    ConstrainedPickupDistEnv,
    ConstrainedGoToObjDoorEnv,
    ConstrainedOpenDoorEnv,
    ConstrainedOpenDoorLocEnv,
    ConstrainedOpenDoorsOrderEnv,
    ConstrainedActionObjDoorEnv,
    ConstrainedGoToOpenEnv,
    ConstrainedFindObjS5Env,
)
from sampler_lang import (
    BabyAIMissionTaskWrapper,
    SentenceMissionEncoder,
    MissionParamAdapter,
    ConstraintParamAdapter,
    preprocess_obs,
)
from maml_rl.policies.categorical_mlp import CategoricalMLPPolicy


@contextmanager
def silence():
    """Suppress noisy BabyAI rejection logs while keeping unexpected exceptions visible."""
    real_print = builtins.print
    buf = io.StringIO()

    def filtered_print(*args, **kwargs):
        if args and isinstance(args[0], str) and "Sampling rejected" in args[0]:
            return
        return real_print(*args, **kwargs)

    builtins.print = filtered_print
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            yield
    finally:
        builtins.print = real_print


def set_seed(seed: int):
    os.environ["PYTHONHASHSEED"] = str(seed)
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


# ---------------------------- Mission definitions ----------------------------
OBJECTS = ["box"]
COLORS = ["red", "green", "blue", "purple", "yellow", "grey"]
PREP_LOCS = ["on", "at", "to"]
LOC_NAMES = ["right", "front"]
DOOR_COLORS = ["yellow", "grey"]

CONSTRAINT_TEXTS = [f"avoid {hazard}" for hazard in HAZARD_TYPES]
DOUBLE_CONSTRAINT_TEXTS = [
    f"avoid {h1} and avoid {h2}" for h1, h2 in itertools.combinations(HAZARD_TYPES.keys(), 2)
]

LOCAL_MISSIONS = [f"go to the {c} {o}" for c in COLORS for o in OBJECTS]
PICKUP_MISSIONS = [f"pick up the {c} {o}" for c in COLORS for o in OBJECTS]
DOOR_MISSIONS = [f"go to the {c} door" for c in DOOR_COLORS]
OPENDOOR_MISSIONS = [f"open the {c} door" for c in DOOR_COLORS]
OPENDOORLOC_MISSIONS = [f"open the door {p} the {l}" for p in PREP_LOCS for l in LOC_NAMES]
OPENDOORSORDER_MISSIONS = (
    [f"open the {c} door" for c in DOOR_COLORS]
    + [f"open the {c1} door, then open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
    + [f"open the {c1} door after you open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
)
ACTIONOBJDOOR_MISSIONS = (
    [f"pick up the {c} {t}" for c in COLORS for t in ["box"]]
    + [f"go to the {c} {t}" for c in COLORS for t in ["box"]]
    + [f"go to the {c} door" for c in DOOR_COLORS]
    + [f"open the {c} door" for c in DOOR_COLORS]
)
FINDOBJS5_MISSIONS = ["pick up the box"]

GOALS_MAP = {
    "ConstrainedGoToLocal": LOCAL_MISSIONS,
    "ConstrainedPickupDist": PICKUP_MISSIONS,
    "ConstrainedGoToObjDoor": LOCAL_MISSIONS + DOOR_MISSIONS,
    "ConstrainedGoToOpen": LOCAL_MISSIONS,
    "ConstrainedOpenDoor": OPENDOOR_MISSIONS,
    "ConstrainedOpenDoorLoc": OPENDOOR_MISSIONS + OPENDOORLOC_MISSIONS,
    "ConstrainedOpenDoorsOrder": OPENDOORSORDER_MISSIONS,
    "ConstrainedActionObjDoor": ACTIONOBJDOOR_MISSIONS,
    "ConstrainedFindObjS5": FINDOBJS5_MISSIONS,
}


def build_env(env_name, room_size, num_dists, max_steps, missions, goals, constraints,
              hazard_density=0.2, max_hazards=4):
    dispatch = {
        "ConstrainedGoToLocal": lambda: ConstrainedGoToLocalEnv(
            room_size=room_size, num_dists=num_dists, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedPickupDist": lambda: ConstrainedPickupDistEnv(
            room_size=room_size, num_dists=num_dists, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedGoToObjDoor": lambda: ConstrainedGoToObjDoorEnv(
            max_steps=max_steps, num_distractors=num_dists,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedGoToOpen": lambda: ConstrainedGoToOpenEnv(
            room_size=room_size, num_dists=num_dists, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedOpenDoor": lambda: ConstrainedOpenDoorEnv(
            room_size=room_size, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedOpenDoorLoc": lambda: ConstrainedOpenDoorLocEnv(
            room_size=room_size, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedOpenDoorsOrder": lambda: ConstrainedOpenDoorsOrderEnv(
            room_size=room_size, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedActionObjDoor": lambda: ConstrainedActionObjDoorEnv(
            room_size=room_size, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedFindObjS5": lambda: ConstrainedFindObjS5Env(
            room_size=5, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
    }
    base = dispatch[env_name]()
    return BabyAIMissionTaskWrapper(base, missions=missions, goals=goals, constraints=constraints)


def make_policy(input_size, output_size, hidden_sizes=(64, 64)):
    return CategoricalMLPPolicy(
        input_size=input_size,
        output_size=output_size,
        hidden_sizes=hidden_sizes,
        nonlinearity=torch.nn.functional.tanh,
    )


def split_goal_constraint(mission: str):
    parts = mission.split(" and avoid ", 1)
    if len(parts) == 2:
        return parts[0], "avoid " + parts[1]
    return mission, None


def discounted_returns(rewards, gamma):
    out = []
    running = 0.0
    for r in reversed(rewards):
        running = r + gamma * running
        out.append(running)
    return list(reversed(out))


def evaluate_policy(env, policy, mission, device, n_episodes, seeds, max_steps, params=None):
    policy.eval()
    ep_steps, ep_success, ep_violations, ep_costs = [], [], [], []
    env_max = getattr(env.unwrapped, "max_steps", max_steps)

    for ep in range(n_episodes):
        env.reset_task(mission)
        with silence():
            obs, _ = env.reset(seed=seeds[ep])
        done = False
        steps = 0
        success = False
        violations = 0
        total_cost = 0.0
        while not done and steps < env_max:
            obs_t = torch.from_numpy(preprocess_obs(obs)[None]).float().to(device)
            with torch.no_grad():
                dist = policy(obs_t, params=params) if params is not None else policy(obs_t)
                action = dist.sample().item()
            obs, reward, terminated, truncated, info = env.step(action)
            done = terminated or truncated
            steps += 1
            c = float(info.get("cost", 0.0))
            total_cost += c
            violations += int(c > 0)
            if terminated:
                success = True

        ep_steps.append(steps)
        ep_success.append(success)
        ep_violations.append(violations)
        ep_costs.append(total_cost)

    return {
        "sr": float(np.mean(ep_success)),
        "steps": float(np.mean(ep_steps)),
        "viols": float(np.mean(ep_violations)),
        "cost": float(np.mean(ep_costs)),
    }


def train_scratch_constraint_rl(env, mission, input_size, output_size, device, args, target=None):
    """Train a fresh policy on one mission using penalized REINFORCE.

    Objective maximized by policy gradient:
        sum_t gamma^t * (reward_t - lambda_cost * cost_t)

    This is deliberately simple and defensible as "Constraint RL from scratch".
    It learns from direct interaction with the target task and does not use meta-learning.
    """
    policy = make_policy(input_size, output_size).to(device)
    optimizer = torch.optim.Adam(policy.parameters(), lr=args.scratch_lr)

    history = []
    best_eval = {"sr": 0.0, "steps": float("inf"), "viols": float("inf"), "cost": float("inf")}
    best_iter = 0
    best_time = 0.0
    total_env_episodes = 0
    start_time = time.time()

    train_seed_rng = random.Random(args.seed + 12345)

    for it in range(1, args.scratch_iters + 1):
        policy.train()
        batch_losses = []
        batch_steps, batch_success, batch_violations, batch_costs = [], [], [], []

        for _ in range(args.scratch_batch_size):
            env.reset_task(mission)
            with silence():
                obs, _ = env.reset(seed=train_seed_rng.randint(0, 10**9))

            done = False
            steps = 0
            rewards_penalized = []
            log_probs = []
            entropies = []
            violations = 0
            total_cost = 0.0
            success = False
            env_max = getattr(env.unwrapped, "max_steps", args.max_steps)

            while not done and steps < env_max:
                obs_t = torch.from_numpy(preprocess_obs(obs)[None]).float().to(device)
                dist = policy(obs_t)
                action_t = dist.sample()
                log_prob = dist.log_prob(action_t)
                entropy = dist.entropy().mean()

                obs, reward, terminated, truncated, info = env.step(int(action_t.item()))
                raw_cost = float(info.get("cost", 0.0))
                shaped_reward = float(reward) - args.lambda_cost * raw_cost

                rewards_penalized.append(shaped_reward)
                log_probs.append(log_prob.squeeze())
                entropies.append(entropy)

                violations += int(raw_cost > 0)
                total_cost += raw_cost
                done = terminated or truncated
                steps += 1
                if terminated:
                    success = True

            returns = torch.tensor(discounted_returns(rewards_penalized, args.gamma), dtype=torch.float32, device=device)
            if returns.numel() > 1:
                returns = (returns - returns.mean()) / (returns.std(unbiased=False) + 1e-8)
            log_probs_t = torch.stack(log_probs)
            entropies_t = torch.stack(entropies)
            loss = -(log_probs_t * returns.detach()).sum() - args.entropy_coef * entropies_t.sum()
            batch_losses.append(loss)

            batch_steps.append(steps)
            batch_success.append(success)
            batch_violations.append(violations)
            batch_costs.append(total_cost)
            total_env_episodes += 1

        optimizer.zero_grad(set_to_none=True)
        loss = torch.stack(batch_losses).mean()
        loss.backward()
        torch.nn.utils.clip_grad_norm_(policy.parameters(), args.max_grad_norm)
        optimizer.step()

        train_row = {
            "iter": it,
            "train_sr": float(np.mean(batch_success)),
            "train_steps": float(np.mean(batch_steps)),
            "train_viols": float(np.mean(batch_violations)),
            "train_cost": float(np.mean(batch_costs)),
            "loss": float(loss.detach().cpu().item()),
        }

        # Evaluate periodically and test whether scratch RL has matched C-LAMAML.
        if it == 1 or it % args.eval_every == 0 or it == args.scratch_iters:
            eval_seeds = [args.seed * 100000 + j for j in range(args.eval_episodes)]
            eval_row = evaluate_policy(env, policy, mission, device, args.eval_episodes, eval_seeds, args.max_steps)
            train_row.update({f"eval_{k}": v for k, v in eval_row.items()})
            history.append(train_row)

            better = (
                eval_row["sr"] > best_eval["sr"]
                or (np.isclose(eval_row["sr"], best_eval["sr"]) and eval_row["steps"] < best_eval["steps"])
            )
            if better:
                best_eval = eval_row
                best_iter = it
                best_time = time.time() - start_time

            if target is not None and matched_target(eval_row, target, args):
                return policy, history, {
                    "matched": True,
                    "match_iter": it,
                    "match_seconds": time.time() - start_time,
                    "env_episodes": total_env_episodes,
                    "best_iter": best_iter,
                    "best_seconds": best_time,
                    **{f"best_{k}": v for k, v in best_eval.items()},
                    **{f"final_{k}": v for k, v in eval_row.items()},
                }
        else:
            history.append(train_row)

    final_eval = evaluate_policy(
        env, policy, mission, device, args.eval_episodes,
        [args.seed * 100000 + j for j in range(args.eval_episodes)], args.max_steps
    )
    return policy, history, {
        "matched": False,
        "match_iter": "",
        "match_seconds": "",
        "env_episodes": total_env_episodes,
        "best_iter": best_iter,
        "best_seconds": best_time,
        **{f"best_{k}": v for k, v in best_eval.items()},
        **{f"final_{k}": v for k, v in final_eval.items()},
    }


def matched_target(eval_row, target, args):
    """Define when scratch Constraint RL has matched C-LAMAML.

    Default rule is intentionally strict but not impossible:
      - success rate must be within sr_tolerance of C-LAMAML
      - average steps must be no worse than C-LAMAML by more than steps_tolerance fraction
      - violations must be no worse than C-LAMAML by more than viol_tolerance absolute count
    """
    return (
        eval_row["sr"] >= max(0.0, target["sr"] - args.sr_tolerance)
        and eval_row["steps"] <= target["steps"] * (1.0 + args.steps_tolerance)
        and eval_row["viols"] <= target["viols"] + args.viol_tolerance
    )


def load_clamaml(args, device, input_size, output_size):
    ckpt_path = f"lang_model/lang_{args.env_name}_{args.delta_theta}_{args.num_constraints}c.pth"
    if not os.path.exists(ckpt_path):
        raise FileNotFoundError(f"C-LAMAML checkpoint not found: {ckpt_path}")

    ckpt = torch.load(ckpt_path, map_location=device)
    policy = make_policy(input_size, output_size).to(device)
    policy.load_state_dict(ckpt["policy"])
    policy.eval()

    policy_param_shapes = [p.shape for p in make_policy(input_size, output_size).parameters()]
    mission_encoder = SentenceMissionEncoder(
        model_name="all-MiniLM-L6-v2",
        frozen=True,
        normalize=True,
        cache=True,
        device=device,
    )
    mission_encoder.eval()
    enc_dim = mission_encoder.output_dim

    mission_adapter = MissionParamAdapter(enc_dim, policy_param_shapes).to(device)
    mission_adapter.load_state_dict(ckpt["mission_adapter"])
    mission_adapter.eval()

    constraint_adapter = None
    if "constraint_adapter" in ckpt:
        constraint_adapter = ConstraintParamAdapter(enc_dim, policy_param_shapes).to(device)
        constraint_adapter.load_state_dict(ckpt["constraint_adapter"])
        constraint_adapter.eval()

    def adapted_params(mission: str):
        goal_str, constr_str = split_goal_constraint(mission)
        with torch.no_grad():
            g_emb = mission_encoder(goal_str).to(device)
            deltas_g = mission_adapter(g_emb)
            deltas_c = [torch.zeros_like(d) for d in deltas_g]
            if constr_str is not None and constraint_adapter is not None:
                c_emb = mission_encoder(constr_str).to(device)
                deltas_c = constraint_adapter(c_emb)

            names = list(dict(policy.named_parameters()).keys())
            params = list(policy.parameters())
            return OrderedDict(
                (n, p + dg.squeeze(0) * args.delta_theta + dc.squeeze(0) * args.delta_constraint)
                for n, p, dg, dc in zip(names, params, deltas_g, deltas_c)
            )

    return policy, adapted_params, ckpt_path


def save_excel(args, rows, aggregate):
    os.makedirs(args.out_dir, exist_ok=True)
    xlsx_path = os.path.join(args.out_dir, args.excel_name)
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet = f"{args.env_name}_{args.num_constraints}c"[:31]
    if sheet in wb.sheetnames:
        del wb[sheet]
    ws = wb.create_sheet(sheet)

    headers = [
        "Mission",
        "C-LAMAML SR%", "C-LAMAML Steps", "C-LAMAML Viols",
        "Constraint RL Matched", "Match Iter", "Match Seconds", "Env Episodes",
        "Best Scratch SR%", "Best Scratch Steps", "Best Scratch Viols",
        "Final Scratch SR%", "Final Scratch Steps", "Final Scratch Viols",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r["mission"],
            round(r["clamaml_sr"] * 100, 2), round(r["clamaml_steps"], 2), round(r["clamaml_viols"], 2),
            r["matched"], r["match_iter"], r["match_seconds"], r["env_episodes"],
            round(r["best_sr"] * 100, 2), round(r["best_steps"], 2), round(r["best_viols"], 2),
            round(r["final_sr"] * 100, 2), round(r["final_steps"], 2), round(r["final_viols"], 2),
        ])

    ws.append([])
    ws.append(["OVERALL"])
    for k, v in aggregate.items():
        ws.append([k, v])

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 55)

    wb.save(xlsx_path)
    return xlsx_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--env", dest="env_name", choices=list(GOALS_MAP.keys()), default="ConstrainedGoToLocal")
    parser.add_argument("--room-size", type=int, default=8)
    parser.add_argument("--num-dists", type=int, default=2)
    parser.add_argument("--max-steps", type=int, default=300)
    parser.add_argument("--delta-theta", type=float, default=0.3)
    parser.add_argument("--delta-constraint", type=float, default=0.1)
    parser.add_argument("--num-constraints", type=int, default=1, choices=[1, 2])
    parser.add_argument("--n-missions", type=int, default=10)
    parser.add_argument("--eval-episodes", type=int, default=10)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--hazard-density", type=float, default=0.2)
    parser.add_argument("--max-hazards", type=int, default=4)

    # Scratch Constraint RL hyperparameters
    parser.add_argument("--scratch-iters", type=int, default=200)
    parser.add_argument("--scratch-batch-size", type=int, default=20)
    parser.add_argument("--scratch-lr", type=float, default=3e-4)
    parser.add_argument("--lambda-cost", type=float, default=1.0)
    parser.add_argument("--gamma", type=float, default=0.99)
    parser.add_argument("--entropy-coef", type=float, default=0.01)
    parser.add_argument("--max-grad-norm", type=float, default=1.0)
    parser.add_argument("--eval-every", type=int, default=10)

    # Matching tolerances
    parser.add_argument("--sr-tolerance", type=float, default=0.05)
    parser.add_argument("--steps-tolerance", type=float, default=0.10)
    parser.add_argument("--viol-tolerance", type=float, default=0.5)

    parser.add_argument("--out-dir", type=str, default="baseline_results")
    parser.add_argument("--excel-name", type=str, default="constraint_rl_vs_clamaml.xlsx")
    parser.add_argument("--save-scratch-policies", action="store_true")
    args = parser.parse_args()

    set_seed(args.seed)
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    goals = GOALS_MAP[args.env_name]
    constraints = CONSTRAINT_TEXTS if args.num_constraints == 1 else DOUBLE_CONSTRAINT_TEXTS
    all_missions = [f"{g} and {c}" for g in goals for c in constraints]

    env = build_env(
        args.env_name, args.room_size, args.num_dists, args.max_steps,
        all_missions, goals, constraints,
        hazard_density=args.hazard_density, max_hazards=args.max_hazards,
    )
    dummy_obs, _ = env.reset()
    input_size = preprocess_obs(dummy_obs).shape[0]
    output_size = env.action_space.n

    clamaml_policy, clamaml_params_fn, ckpt_path = load_clamaml(args, device, input_size, output_size)

    rng = random.Random(args.seed)
    test_tasks = rng.sample(all_missions, min(args.n_missions, len(all_missions)))
    eval_seeds = [args.seed * 100000 + j for j in range(args.eval_episodes)]

    print("\n" + "=" * 80)
    print("C-LAMAML vs Constraint RL Scratch Baseline")
    print(f"Env: {args.env_name} | Tasks: {len(test_tasks)} | Eval episodes/task: {args.eval_episodes}")
    print(f"Loaded C-LAMAML: {ckpt_path}")
    print("Constraint RL is trained from random initialization separately for each unseen task.")
    print("=" * 80)

    rows = []
    os.makedirs(args.out_dir, exist_ok=True)
    if args.save_scratch_policies:
        os.makedirs(os.path.join(args.out_dir, "scratch_policies"), exist_ok=True)

    for idx, mission in enumerate(test_tasks, 1):
        print(f"\n[{idx}/{len(test_tasks)}] Mission: {mission}")

        env.reset_task(mission)
        c_params = clamaml_params_fn(mission)
        clamaml_eval = evaluate_policy(
            env, clamaml_policy, mission, device, args.eval_episodes, eval_seeds, args.max_steps, params=c_params
        )
        print(
            f"  C-LAMAML target: SR={clamaml_eval['sr']*100:.1f}% | "
            f"Steps={clamaml_eval['steps']:.1f} | Viols={clamaml_eval['viols']:.2f}"
        )

        scratch_policy, hist, scratch_summary = train_scratch_constraint_rl(
            env, mission, input_size, output_size, device, args, target=clamaml_eval
        )

        safe_name = mission.replace(" ", "_").replace(",", "").replace("/", "_")[:80]
        hist_path = os.path.join(args.out_dir, f"history_{idx:02d}_{safe_name}.json")
        with open(hist_path, "w") as f:
            json.dump(hist, f, indent=2)

        if args.save_scratch_policies:
            torch.save(
                scratch_policy.state_dict(),
                os.path.join(args.out_dir, "scratch_policies", f"scratch_{idx:02d}_{safe_name}.pth"),
            )

        matched_text = "YES" if scratch_summary["matched"] else "NO"
        print(
            f"  Constraint RL: matched={matched_text} | "
            f"match_iter={scratch_summary['match_iter']} | "
            f"episodes={scratch_summary['env_episodes']} | "
            f"best_SR={scratch_summary['best_sr']*100:.1f}% | "
            f"best_steps={scratch_summary['best_steps']:.1f} | "
            f"best_viols={scratch_summary['best_viols']:.2f}"
        )

        rows.append({
            "mission": mission,
            "clamaml_sr": clamaml_eval["sr"],
            "clamaml_steps": clamaml_eval["steps"],
            "clamaml_viols": clamaml_eval["viols"],
            "matched": matched_text,
            "match_iter": scratch_summary["match_iter"],
            "match_seconds": scratch_summary["match_seconds"],
            "env_episodes": scratch_summary["env_episodes"],
            "best_sr": scratch_summary["best_sr"],
            "best_steps": scratch_summary["best_steps"],
            "best_viols": scratch_summary["best_viols"],
            "final_sr": scratch_summary["final_sr"],
            "final_steps": scratch_summary["final_steps"],
            "final_viols": scratch_summary["final_viols"],
        })

    matched = [r for r in rows if r["matched"] == "YES"]
    aggregate = {
        "Matched tasks": f"{len(matched)}/{len(rows)}",
        "Avg C-LAMAML SR%": round(float(np.mean([r["clamaml_sr"] for r in rows])) * 100, 2),
        "Avg C-LAMAML steps": round(float(np.mean([r["clamaml_steps"] for r in rows])), 2),
        "Avg C-LAMAML violations": round(float(np.mean([r["clamaml_viols"] for r in rows])), 2),
        "Avg best Constraint RL SR%": round(float(np.mean([r["best_sr"] for r in rows])) * 100, 2),
        "Avg best Constraint RL steps": round(float(np.mean([r["best_steps"] for r in rows])), 2),
        "Avg best Constraint RL violations": round(float(np.mean([r["best_viols"] for r in rows])), 2),
        "Avg match iterations on matched tasks": round(float(np.mean([r["match_iter"] for r in matched])), 2) if matched else "NA",
        "Avg match seconds on matched tasks": round(float(np.mean([r["match_seconds"] for r in matched])), 2) if matched else "NA",
        "Avg env episodes used": round(float(np.mean([r["env_episodes"] for r in rows])), 2),
    }

    xlsx_path = save_excel(args, rows, aggregate)

    print("\n" + "=" * 80)
    print("Overall comparison")
    for k, v in aggregate.items():
        print(f"  {k}: {v}")
    print(f"Results saved to: {xlsx_path}")
    print("=" * 80)


if __name__ == "__main__":
    main()
