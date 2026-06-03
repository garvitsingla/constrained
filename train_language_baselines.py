import os
import time
import json
import random
import argparse
import warnings
import logging
from collections import OrderedDict
from contextlib import contextmanager, redirect_stdout, redirect_stderr
import builtins
import io
import itertools

import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.nn.utils.convert_parameters import parameters_to_vector, vector_to_parameters
from torch.distributions.kl import kl_divergence
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

warnings.filterwarnings("ignore")
logging.getLogger("gymnasium").setLevel(logging.ERROR)
os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")

from environment import (
    HAZARD_TYPES,
    ConstrainedGoToLocalEnv,
    ConstrainedPickupDistEnv,
    ConstrainedGoToObjDoorEnv,
    ConstrainedOpenDoorEnv,
    ConstrainedOpenDoorLocEnv,
    ConstrainedOpenDoorsOrderEnv,
    ConstrainedActionObjDoorEnv,
    ConstrainedGoToOpenEnv,
    ConstrainedFindObjS5Env,
)
from sampler_lang import (
    BabyAIMissionTaskWrapper,
    SentenceMissionEncoder,
    MissionParamAdapter,
    ConstraintParamAdapter,
    preprocess_obs,
)
from maml_rl.policies.categorical_mlp import CategoricalMLPPolicy
from torch.distributions import Categorical
from maml_rl.metalearners.lang_trpo import MAMLTRPO
from maml_rl.baseline import LinearFeatureBaseline
from maml_rl.episode import BatchEpisodes
from maml_rl.utils.torch_utils import weighted_mean, detach_distribution
from maml_rl.utils.optimization import conjugate_gradient


@contextmanager
def silence():
    real_print = builtins.print
    buf = io.StringIO()
    def filtered_print(*args, **kwargs):
        if args and isinstance(args[0], str) and "Sampling rejected" in args[0]:
            return
        return real_print(*args, **kwargs)
    builtins.print = filtered_print
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            yield
    finally:
        builtins.print = real_print

def set_seed(seed: int):
    os.environ["PYTHONHASHSEED"] = str(seed)
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

# ---------------------------- Mission definitions ----------------------------
OBJECTS = ["box"]
COLORS = ["red", "green", "blue", "purple", "yellow", "grey"]
PREP_LOCS = ["on", "at", "to"]
LOC_NAMES = ["right", "front"]
DOOR_COLORS = ["yellow", "grey"]

CONSTRAINT_TEXTS = [f"avoid {hazard}" for hazard in HAZARD_TYPES]
DOUBLE_CONSTRAINT_TEXTS = [
    f"avoid {h1} and avoid {h2}" for h1, h2 in itertools.combinations(HAZARD_TYPES.keys(), 2)
]

LOCAL_MISSIONS = [f"go to the {c} {o}" for c in COLORS for o in OBJECTS]
PICKUP_MISSIONS = [f"pick up the {c} {o}" for c in COLORS for o in OBJECTS]
DOOR_MISSIONS = [f"go to the {c} door" for c in DOOR_COLORS]
OPENDOOR_MISSIONS = [f"open the {c} door" for c in DOOR_COLORS]
OPENDOORLOC_MISSIONS = [f"open the door {p} the {l}" for p in PREP_LOCS for l in LOC_NAMES]
OPENDOORSORDER_MISSIONS = (
    [f"open the {c} door" for c in DOOR_COLORS]
    + [f"open the {c1} door, then open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
    + [f"open the {c1} door after you open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
)
ACTIONOBJDOOR_MISSIONS = (
    [f"pick up the {c} {t}" for c in COLORS for t in ["box"]]
    + [f"go to the {c} {t}" for c in COLORS for t in ["box"]]
    + [f"go to the {c} door" for c in DOOR_COLORS]
    + [f"open the {c} door" for c in DOOR_COLORS]
)
FINDOBJS5_MISSIONS = ["pick up the box"]

GOALS_MAP = {
    "ConstrainedGoToLocal": LOCAL_MISSIONS,
    "ConstrainedPickupDist": PICKUP_MISSIONS,
    "ConstrainedGoToObjDoor": LOCAL_MISSIONS + DOOR_MISSIONS,
    "ConstrainedGoToOpen": LOCAL_MISSIONS,
    "ConstrainedOpenDoor": OPENDOOR_MISSIONS,
    "ConstrainedOpenDoorLoc": OPENDOOR_MISSIONS + OPENDOORLOC_MISSIONS,
    "ConstrainedOpenDoorsOrder": OPENDOORSORDER_MISSIONS,
    "ConstrainedActionObjDoor": ACTIONOBJDOOR_MISSIONS,
    "ConstrainedFindObjS5": FINDOBJS5_MISSIONS,
}

def build_env(env_name, room_size, num_dists, max_steps, missions, goals, constraints,
              hazard_density=0.2, max_hazards=4):
    dispatch = {
        "ConstrainedGoToLocal": lambda: ConstrainedGoToLocalEnv(
            room_size=room_size, num_dists=num_dists, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedPickupDist": lambda: ConstrainedPickupDistEnv(
            room_size=room_size, num_dists=num_dists, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedGoToObjDoor": lambda: ConstrainedGoToObjDoorEnv(
            max_steps=max_steps, num_distractors=num_dists,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedGoToOpen": lambda: ConstrainedGoToOpenEnv(
            room_size=room_size, num_dists=num_dists, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedOpenDoor": lambda: ConstrainedOpenDoorEnv(
            room_size=room_size, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedOpenDoorLoc": lambda: ConstrainedOpenDoorLocEnv(
            room_size=room_size, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedOpenDoorsOrder": lambda: ConstrainedOpenDoorsOrderEnv(
            room_size=room_size, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedActionObjDoor": lambda: ConstrainedActionObjDoorEnv(
            room_size=room_size, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
        "ConstrainedFindObjS5": lambda: ConstrainedFindObjS5Env(
            room_size=5, max_steps=max_steps,
            hazard_density=hazard_density, max_hazards=max_hazards),
    }
    base = dispatch[env_name]()
    return BabyAIMissionTaskWrapper(base, missions=missions, goals=goals, constraints=constraints)

def make_policy(input_size, output_size, hidden_sizes=(64, 64)):
    return CategoricalMLPPolicy(
        input_size=input_size,
        output_size=output_size,
        hidden_sizes=hidden_sizes,
        nonlinearity=torch.nn.functional.tanh,
    )

class ConcatMLPPolicy(nn.Module):
    def __init__(self, input_size, output_size, hidden_sizes=(64, 64)):
        super(ConcatMLPPolicy, self).__init__()
        self.fc1 = nn.Linear(input_size, hidden_sizes[0])
        self.fc2 = nn.Linear(hidden_sizes[0], hidden_sizes[1])
        self.out = nn.Linear(hidden_sizes[1], output_size)
    
    def forward(self, x):
        x = torch.tanh(self.fc1(x))
        x = torch.tanh(self.fc2(x))
        logits = self.out(x)
        return Categorical(logits=logits)

def split_goal_constraint(mission: str):
    parts = mission.split(" and avoid ", 1)
    if len(parts) == 2:
        return parts[0], "avoid " + parts[1]
    return mission, None

class StandardTRPO:
    def __init__(self, policy, device='cpu'):
        self.policy = policy
        self.device = device
        
    def hessian_vector_product(self, kl, meta_params, damping=1e-2):
        grads = torch.autograd.grad(kl, meta_params, create_graph=True)
        flat_grad_kl = parameters_to_vector(grads)
        def _product(vector, retain_graph=True):
            grad_kl_v = torch.dot(flat_grad_kl, vector)
            grad2s = torch.autograd.grad(grad_kl_v, meta_params, retain_graph=retain_graph)
            flat_grad2_kl = parameters_to_vector(grad2s)
            return flat_grad2_kl + damping * vector
        return _product

    def surrogate_loss(self, episodes, old_pi=None):
        with torch.set_grad_enabled(old_pi is None):
            pi = self.policy(episodes.observations)
            if old_pi is None:
                old_pi_task = detach_distribution(pi)
            else:
                old_pi_task = old_pi

            if isinstance(old_pi_task, list):
                old_log_prob = old_pi_task[0].log_prob(episodes.actions)
            else:
                old_log_prob = old_pi_task.log_prob(episodes.actions)
                
            log_ratio = pi.log_prob(episodes.actions) - old_log_prob
            ratio = torch.exp(log_ratio)
            
            if hasattr(episodes, '_cost_advantages') and episodes._cost_advantages is not None:
                combined_adv = episodes.advantages - episodes.cost_advantages
            else:
                combined_adv = episodes.advantages
                
            loss = -weighted_mean(ratio * combined_adv, lengths=episodes.lengths)
            kl = weighted_mean(kl_divergence(pi, old_pi), lengths=episodes.lengths)
        return loss.mean(), kl.mean(), old_pi

    def step(self, episodes, max_kl=1e-3, cg_iters=10, cg_damping=1e-2, ls_max_steps=10, ls_backtrack_ratio=0.5):
        old_loss, old_kl, old_pi = self.surrogate_loss(episodes, old_pi=None)
        
        meta_params = [p for p in self.policy.parameters() if p.requires_grad]
        grads = torch.autograd.grad(old_loss, meta_params, retain_graph=True)
        grads = parameters_to_vector(grads)
        
        hessian_vector_product = self.hessian_vector_product(old_kl, meta_params=meta_params, damping=cg_damping)
        stepdir = conjugate_gradient(hessian_vector_product, grads, cg_iters=cg_iters)
        
        shs = 0.5 * torch.dot(stepdir, hessian_vector_product(stepdir, retain_graph=False))
        lagrange_multiplier = torch.sqrt(shs / max_kl)
        step = stepdir / lagrange_multiplier
        
        old_params = parameters_to_vector(meta_params)
        
        step_size = 1.0
        for _ in range(ls_max_steps):
            vector_to_parameters(old_params - step_size * step, meta_params)
            loss, kl, _ = self.surrogate_loss(episodes, old_pi=old_pi)
            improve = loss - old_loss
            if (improve.item() < 0.0) and (kl.item() < max_kl):
                break
            step_size *= ls_backtrack_ratio
        else:
            vector_to_parameters(old_params, meta_params)

def discounted_returns(rewards, gamma):
    out = []
    running = 0.0
    for r in reversed(rewards):
        running = r + gamma * running
        out.append(running)
    return list(reversed(out))

# ==================== ADAPTER BASELINE ====================

def get_adapted_params(policy, encoder, m_adapter, c_adapter, mission, args, device):
    goal_str, constr_str = split_goal_constraint(mission)
    g_emb = encoder(goal_str).to(device)
    deltas_g = m_adapter(g_emb)
    deltas_c = [torch.zeros_like(d) for d in deltas_g]
    if constr_str is not None and c_adapter is not None:
        c_emb = encoder(constr_str).to(device)
        deltas_c = c_adapter(c_emb)

    names = list(dict(policy.named_parameters()).keys())
    params = list(policy.parameters())
    
    return OrderedDict(
        (n, p + dg.squeeze(0) * args.delta_theta + dc.squeeze(0) * args.delta_constraint)
        for n, p, dg, dc in zip(names, params, deltas_g, deltas_c)
    )

def train_adapter_baseline(env, mission, input_size, output_size, device, args, target=None):
    policy = make_policy(input_size, output_size).to(device)
    policy_param_shapes = [p.shape for p in policy.parameters()]
    baseline = LinearFeatureBaseline(input_size).to(device)
    cost_baseline = LinearFeatureBaseline(input_size).to(device)
    
    encoder = SentenceMissionEncoder(model_name="all-MiniLM-L6-v2", frozen=True, normalize=True, cache=True, device=device)
    encoder.eval()
    
    m_adapter = MissionParamAdapter(encoder.output_dim, policy_param_shapes).to(device)
    c_adapter = ConstraintParamAdapter(encoder.output_dim, policy_param_shapes).to(device)
    
    lambda_weights = {2: args.lambda_cost, 3: args.lambda_cost, 4: args.lambda_cost}
    metalearner = MAMLTRPO(
        policy=policy,
        mission_encoder=encoder,
        mission_adapter=m_adapter,
        constraint_adapter=c_adapter,
        delta_theta=args.delta_theta,
        delta_constraint=args.delta_constraint,
        first_order=True,
        device=device,
        lambda_weights=lambda_weights
    )

    best_eval = {"sr": 0.0, "steps": float("inf"), "viols": float("inf")}
    total_env_episodes = 0
    start_time = time.time()
    train_seed_rng = random.Random(args.seed + 12345)

    HAZARD_TYPES = {'lava': 2, 'grass': 3, 'water': 4}
    constraint_tiles = []
    mission_str = split_goal_constraint(mission)[1] or ""
    for hazard, idx in HAZARD_TYPES.items():
        if f"avoid {hazard}" in str(mission_str):
            constraint_tiles.append(idx)

    for it in range(1, args.scratch_iters + 1):
        policy.train(); m_adapter.train(); c_adapter.train()

        batch_episodes = BatchEpisodes(batch_size=args.scratch_batch_size, gamma=args.gamma, device=device)
        batch_episodes.mission = split_goal_constraint(mission)
        batch_episodes.constraint_tiles = constraint_tiles

        adapted_params = metalearner.adapt_one(split_goal_constraint(mission))

        for ep in range(args.scratch_batch_size):
            env.reset_task(mission)
            with silence():
                obs, _ = env.reset(seed=train_seed_rng.randint(0, 10**9))
            done = False
            steps = 0
            env_max = getattr(env.unwrapped, "max_steps", args.max_steps)

            while not done and steps < env_max:
                obs_vec = preprocess_obs(obs)
                obs_t = torch.from_numpy(obs_vec[None]).float().to(device)
                
                with torch.no_grad():
                    dist = policy(obs_t, params=adapted_params)
                    action = dist.sample().item()

                obs, reward, terminated, truncated, info = env.step(action)
                cost = float(info.get("cost", 0.0))
                
                # Apply lambda weighting during sampling exactly like sampler_lang.py
                tile_idx = info.get('tile_index', 0)
                if cost > 0:
                    cost = cost * lambda_weights.get(tile_idx, 1.0)
                
                done = terminated or truncated
                steps += 1
                
                batch_episodes.append([obs_vec], [np.array(action)], [np.array(reward)], [np.array(cost)], [np.array(ep)])

        baseline.fit(batch_episodes)
        cost_baseline.fit_costs(batch_episodes)
        batch_episodes.compute_advantages(baseline, gae_lambda=1.0, normalize=True)
        batch_episodes.compute_cost_advantages(cost_baseline, gae_lambda=1.0, normalize=True)
        
        metalearner.step([batch_episodes], [batch_episodes])

        total_env_episodes += args.scratch_batch_size

        # Evaluate
        if it == 1 or it % args.eval_every == 0 or it == args.scratch_iters:
            policy.eval(); m_adapter.eval(); c_adapter.eval()
            eval_params = metalearner.adapt_one(split_goal_constraint(mission))
            eval_res = evaluate_policy(env, policy, mission, device, args.eval_episodes, [args.seed * 1000 + j for j in range(args.eval_episodes)], args.max_steps, eval_params)
            
            print(f"        [Iter {it:03d}] Eval SR={eval_res['sr']*100:5.1f}% | Steps={eval_res['steps']:5.1f} | Viols={eval_res['viols']:5.2f}")
            
            if eval_res["sr"] > best_eval["sr"] or (np.isclose(eval_res["sr"], best_eval["sr"]) and eval_res["steps"] < best_eval["steps"]):
                best_eval = eval_res

            if target is not None and matched_target(eval_res, target, args):
                return {"matched": True, "match_iter": it, "match_seconds": time.time() - start_time, "env_episodes": total_env_episodes, "best_sr": best_eval["sr"], "best_steps": best_eval["steps"], "best_viols": best_eval["viols"]}
    
    return {"matched": False, "match_iter": "", "match_seconds": "", "env_episodes": total_env_episodes, "best_sr": best_eval["sr"], "best_steps": best_eval["steps"], "best_viols": best_eval["viols"]}


# ==================== CONCAT BASELINE ====================

def train_concat_baseline(env, mission, input_size, output_size, device, args, target=None):
    encoder = SentenceMissionEncoder(model_name="all-MiniLM-L6-v2", frozen=True, normalize=True, cache=True, device=device)
    encoder.eval()
    
    # Pre-compute embedding
    goal_str, constr_str = split_goal_constraint(mission)
    g_emb = encoder(goal_str).to(device)
    c_emb = encoder(constr_str).to(device) if constr_str else torch.zeros_like(g_emb)
    combined_emb = torch.cat([g_emb, c_emb], dim=-1) # Shape: (768)
    
    policy = ConcatMLPPolicy(input_size + combined_emb.shape[-1], output_size).to(device)
    baseline = LinearFeatureBaseline(input_size + combined_emb.shape[-1]).to(device)
    cost_baseline = LinearFeatureBaseline(input_size + combined_emb.shape[-1]).to(device)
    
    metalearner = StandardTRPO(policy, device=device)

    best_eval = {"sr": 0.0, "steps": float("inf"), "viols": float("inf")}
    total_env_episodes = 0
    start_time = time.time()
    train_seed_rng = random.Random(args.seed + 12345)

    HAZARD_TYPES = {'lava': 2, 'grass': 3, 'water': 4}
    constraint_tiles = []
    mission_str = constr_str or ""
    for hazard, idx in HAZARD_TYPES.items():
        if f"avoid {hazard}" in mission_str:
            constraint_tiles.append(idx)
    lambda_weights = {2: args.lambda_cost, 3: args.lambda_cost, 4: args.lambda_cost}

    for it in range(1, args.scratch_iters + 1):
        policy.train()

        batch_episodes = BatchEpisodes(batch_size=args.scratch_batch_size, gamma=args.gamma, device=device)
        batch_episodes.mission = mission
        batch_episodes.constraint_tiles = constraint_tiles

        for ep in range(args.scratch_batch_size):
            env.reset_task(mission)
            with silence():
                obs, _ = env.reset(seed=train_seed_rng.randint(0, 10**9))
            done = False
            steps = 0
            env_max = getattr(env.unwrapped, "max_steps", args.max_steps)

            while not done and steps < env_max:
                obs_vec = preprocess_obs(obs)
                obs_t = torch.from_numpy(obs_vec[None]).float().to(device)
                full_input = torch.cat([obs_t, combined_emb], dim=-1) # 1 x (obs + emb)
                
                with torch.no_grad():
                    dist = policy(full_input)
                    action = dist.sample().item()

                obs, reward, terminated, truncated, info = env.step(action)
                cost = float(info.get("cost", 0.0))
                
                tile_idx = info.get('tile_index', 0)
                if cost > 0:
                    cost = cost * lambda_weights.get(tile_idx, 1.0)
                
                done = terminated or truncated
                steps += 1
                
                # Note: We append full_input so the baseline can fit on the concatenated obs!
                batch_episodes.append([full_input.cpu().numpy().squeeze()], [np.array(action)], [np.array(reward)], [np.array(cost)], [np.array(ep)])

        baseline.fit(batch_episodes)
        cost_baseline.fit_costs(batch_episodes)
        batch_episodes.compute_advantages(baseline, gae_lambda=1.0, normalize=True)
        batch_episodes.compute_cost_advantages(cost_baseline, gae_lambda=1.0, normalize=True)
        
        metalearner.step(batch_episodes)
        total_env_episodes += args.scratch_batch_size

        # Evaluate
        if it == 1 or it % args.eval_every == 0 or it == args.scratch_iters:
            policy.eval()
            
            eval_steps, eval_success, eval_violations = [], [], []
            for ep in range(args.eval_episodes):
                env.reset_task(mission)
                with silence():
                    obs, _ = env.reset(seed=(args.seed * 1000 + ep))
                edone = False; esteps = 0; eviolations = 0; esuccess = False
                while not edone and esteps < env_max:
                    obs_t = torch.from_numpy(preprocess_obs(obs)[None]).float().to(device)
                    dist = policy(torch.cat([obs_t, combined_emb], dim=-1))
                    action = dist.sample().item()
                    obs, r, term, trunc, info = env.step(action)
                    edone = term or trunc
                    esteps += 1
                    eviolations += int(info.get("cost", 0.0) > 0)
                    if term: esuccess = True
                eval_steps.append(esteps); eval_success.append(esuccess); eval_violations.append(eviolations)
            
            eval_res = {"sr": float(np.mean(eval_success)), "steps": float(np.mean(eval_steps)), "viols": float(np.mean(eval_violations))}
            
            print(f"        [Iter {it:03d}] Eval SR={eval_res['sr']*100:5.1f}% | Steps={eval_res['steps']:5.1f} | Viols={eval_res['viols']:5.2f}")

            if eval_res["sr"] > best_eval["sr"] or (np.isclose(eval_res["sr"], best_eval["sr"]) and eval_res["steps"] < best_eval["steps"]):
                best_eval = eval_res

            if target is not None and matched_target(eval_res, target, args):
                return {"matched": True, "match_iter": it, "match_seconds": time.time() - start_time, "env_episodes": total_env_episodes, "best_sr": best_eval["sr"], "best_steps": best_eval["steps"], "best_viols": best_eval["viols"]}
    
    return {"matched": False, "match_iter": "", "match_seconds": "", "env_episodes": total_env_episodes, "best_sr": best_eval["sr"], "best_steps": best_eval["steps"], "best_viols": best_eval["viols"]}

# ==================== HELPERS ====================

def evaluate_policy(env, policy, mission, device, n_episodes, seeds, max_steps, params=None):
    policy.eval()
    ep_steps, ep_success, ep_violations = [], [], []
    env_max = getattr(env.unwrapped, "max_steps", max_steps)

    for ep in range(n_episodes):
        env.reset_task(mission)
        with silence():
            obs, _ = env.reset(seed=seeds[ep])
        done = False
        steps = 0
        success = False
        violations = 0
        while not done and steps < env_max:
            obs_t = torch.from_numpy(preprocess_obs(obs)[None]).float().to(device)
            with torch.no_grad():
                dist = policy(obs_t, params=params) if params is not None else policy(obs_t)
                action = dist.sample().item()
            obs, reward, terminated, truncated, info = env.step(action)
            done = terminated or truncated
            steps += 1
            violations += int(info.get("cost", 0) > 0)
            if terminated:
                success = True

        ep_steps.append(steps)
        ep_success.append(success)
        ep_violations.append(violations)

    return {
        "sr": float(np.mean(ep_success)),
        "steps": float(np.mean(ep_steps)),
        "viols": float(np.mean(ep_violations)),
    }

def matched_target(eval_row, target, args):
    return (
        eval_row["sr"] >= max(0.0, target["sr"] - args.sr_tolerance)
        and eval_row["steps"] <= target["steps"] * (1.0 + args.steps_tolerance)
        and eval_row["viols"] <= target["viols"] + args.viol_tolerance
    )

def save_excel(args, rows, aggregate):
    os.makedirs(args.out_dir, exist_ok=True)
    xlsx_path = os.path.join(args.out_dir, args.excel_name)
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet = f"{args.env_name}_{args.num_constraints}c"[:31]
    if sheet in wb.sheetnames:
        del wb[sheet]
    ws = wb.create_sheet(sheet)

    headers = [
        "Mission",
        "C-LAMAML SR%", "C-LAMAML Steps", "C-LAMAML Viols",
        "Base 1 (Adapter) Matched", "Base 1 Iter", "Base 1 Best SR%", "Base 1 Best Steps", "Base 1 Best Viols",
        "Base 2 (Concat) Matched", "Base 2 Iter", "Base 2 Best SR%", "Base 2 Best Steps", "Base 2 Best Viols",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r["mission"],
            round(r["clamaml_sr"] * 100, 2), round(r["clamaml_steps"], 2), round(r["clamaml_viols"], 2),
            r["b1_matched"], r["b1_iter"], round(r["b1_sr"] * 100, 2), round(r["b1_steps"], 2), round(r["b1_viols"], 2),
            r["b2_matched"], r["b2_iter"], round(r["b2_sr"] * 100, 2), round(r["b2_steps"], 2), round(r["b2_viols"], 2),
        ])

    ws.append([])
    ws.append(["OVERALL"])
    for k, v in aggregate.items():
        ws.append([k, v])

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 55)

    wb.save(xlsx_path)
    return xlsx_path

def load_clamaml(args, device, input_size, output_size):
    ckpt_path = f"lang_model/lang_{args.env_name}_{args.delta_theta}_{args.num_constraints}c.pth"
    if not os.path.exists(ckpt_path):
        raise FileNotFoundError(f"C-LAMAML checkpoint not found: {ckpt_path}")

    ckpt = torch.load(ckpt_path, map_location=device)
    policy = make_policy(input_size, output_size).to(device)
    policy.load_state_dict(ckpt["policy"])
    policy.eval()

    policy_param_shapes = [p.shape for p in make_policy(input_size, output_size).parameters()]
    mission_encoder = SentenceMissionEncoder(
        model_name="all-MiniLM-L6-v2", frozen=True, normalize=True, cache=True, device=device
    )
    mission_encoder.eval()
    enc_dim = mission_encoder.output_dim

    mission_adapter = MissionParamAdapter(enc_dim, policy_param_shapes).to(device)
    mission_adapter.load_state_dict(ckpt["mission_adapter"])
    mission_adapter.eval()

    constraint_adapter = None
    if "constraint_adapter" in ckpt:
        constraint_adapter = ConstraintParamAdapter(enc_dim, policy_param_shapes).to(device)
        constraint_adapter.load_state_dict(ckpt["constraint_adapter"])
        constraint_adapter.eval()

    def adapted_params(mission: str):
        goal_str, constr_str = split_goal_constraint(mission)
        with torch.no_grad():
            g_emb = mission_encoder(goal_str).to(device)
            deltas_g = mission_adapter(g_emb)
            deltas_c = [torch.zeros_like(d) for d in deltas_g]
            if constr_str is not None and constraint_adapter is not None:
                c_emb = mission_encoder(constr_str).to(device)
                deltas_c = constraint_adapter(c_emb)

            names = list(dict(policy.named_parameters()).keys())
            params = list(policy.parameters())
            return OrderedDict(
                (n, p + dg.squeeze(0) * args.delta_theta + dc.squeeze(0) * args.delta_constraint)
                for n, p, dg, dc in zip(names, params, deltas_g, deltas_c)
            )

    return policy, adapted_params, ckpt_path

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--env", dest="env_name", choices=list(GOALS_MAP.keys()), default="ConstrainedGoToLocal")
    parser.add_argument("--room-size", type=int, default=8)
    parser.add_argument("--num-dists", type=int, default=2)
    parser.add_argument("--max-steps", type=int, default=300)
    parser.add_argument("--delta-theta", type=float, default=0.3)
    parser.add_argument("--delta-constraint", type=float, default=0.1)
    parser.add_argument("--num-constraints", type=int, default=1)
    parser.add_argument("--n-missions", type=int, default=10)
    parser.add_argument("--eval-episodes", type=int, default=10)
    parser.add_argument("--seed", type=int, default=1)
    parser.add_argument("--hazard-density", type=float, default=0.2)
    parser.add_argument("--max-hazards", type=int, default=4)

    # Scratch RL hyperparameters
    parser.add_argument("--scratch-iters", type=int, default=200)
    parser.add_argument("--scratch-batch-size", type=int, default=40)
    parser.add_argument("--scratch-lr", type=float, default=3e-4)
    parser.add_argument("--lambda-cost", type=float, default=1.0)
    parser.add_argument("--gamma", type=float, default=0.99)
    parser.add_argument("--entropy-coef", type=float, default=0.01)
    parser.add_argument("--max-grad-norm", type=float, default=1.0)
    parser.add_argument("--eval-every", type=int, default=10)

    # Matching tolerances
    parser.add_argument("--sr-tolerance", type=float, default=0.05)
    parser.add_argument("--steps-tolerance", type=float, default=0.10)
    parser.add_argument("--viol-tolerance", type=float, default=0.5)
    
    parser.add_argument("--out-dir", type=str, default="baseline_results")
    parser.add_argument("--excel-name", type=str, default="language_baselines_vs_clamaml.xlsx")

    args = parser.parse_args()
    
    env_name = args.env_name
    room_size = args.room_size
    num_dists = args.num_dists
    max_steps = args.max_steps
    n_missions = args.n_missions
    delta_theta = args.delta_theta
    delta_constraint = args.delta_constraint
    eval_episodes = args.eval_episodes

    set_seed(args.seed)
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    goals = GOALS_MAP[args.env_name]
    constraints = CONSTRAINT_TEXTS if args.num_constraints == 1 else DOUBLE_CONSTRAINT_TEXTS
    all_missions = [f"{g} and {c}" for g in goals for c in constraints]

    env = build_env(
        args.env_name, args.room_size, args.num_dists, args.max_steps,
        all_missions, goals, constraints,
        hazard_density=args.hazard_density, max_hazards=args.max_hazards,
    )
    dummy_obs, _ = env.reset()
    input_size = preprocess_obs(dummy_obs).shape[0]
    output_size = env.action_space.n

    clamaml_policy, clamaml_params_fn, ckpt_path = load_clamaml(args, device, input_size, output_size)

    rng = random.Random(args.seed)
    test_tasks = rng.sample(all_missions, min(args.n_missions, len(all_missions)))
    eval_seeds = [args.seed * 100000 + j for j in range(args.eval_episodes)]

    print("\n" + "=" * 80)
    print("C-LAMAML vs Language-Conditioned Baselines from Scratch")
    print(f"Env: {args.env_name} | Tasks: {len(test_tasks)} | Eval episodes/task: {args.eval_episodes}")
    print("=" * 80)

    rows = []

    for idx, mission in enumerate(test_tasks, 1):
        print(f"\n[{idx}/{len(test_tasks)}] Mission: {mission}")

        # 1. C-LAMAML Zero-Shot Target
        env.reset_task(mission)
        c_params = clamaml_params_fn(mission)
        clamaml_eval = evaluate_policy(env, clamaml_policy, mission, device, args.eval_episodes, eval_seeds, args.max_steps, params=c_params)
        print(f"  [TARGET] C-LAMAML : SR={clamaml_eval['sr']*100:.1f}% | Steps={clamaml_eval['steps']:.1f} | Viols={clamaml_eval['viols']:.2f}")

        # 2. Baseline 1: Adapter from scratch
        print("  --> Training Baseline 1 (Adapter from Scratch)...")
        res1 = train_adapter_baseline(env, mission, input_size, output_size, device, args, target=clamaml_eval)
        print(f"      Matched={res1['matched']} | Iter={res1['match_iter']} | Eps={res1['env_episodes']} | SR={res1['best_sr']*100:.1f}% | Viols={res1['best_viols']:.2f}")

        # 3. Baseline 2: Concat from scratch
        print("  --> Training Baseline 2 (Concat from Scratch)...")
        res2 = train_concat_baseline(env, mission, input_size, output_size, device, args, target=clamaml_eval)
        print(f"      Matched={res2['matched']} | Iter={res2['match_iter']} | Eps={res2['env_episodes']} | SR={res2['best_sr']*100:.1f}% | Viols={res2['best_viols']:.2f}")
        
        rows.append({
            "mission": mission,
            "clamaml_sr": clamaml_eval["sr"],
            "clamaml_steps": clamaml_eval["steps"],
            "clamaml_viols": clamaml_eval["viols"],
            "b1_matched": "YES" if res1["matched"] else "NO",
            "b1_iter": res1["match_iter"],
            "b1_sr": res1["best_sr"],
            "b1_steps": res1["best_steps"],
            "b1_viols": res1["best_viols"],
            "b2_matched": "YES" if res2["matched"] else "NO",
            "b2_iter": res2["match_iter"],
            "b2_sr": res2["best_sr"],
            "b2_steps": res2["best_steps"],
            "b2_viols": res2["best_viols"],
        })

    aggregate = {
        "Base 1 Matched tasks": f"{len([r for r in rows if r['b1_matched'] == 'YES'])}/{len(rows)}",
        "Base 2 Matched tasks": f"{len([r for r in rows if r['b2_matched'] == 'YES'])}/{len(rows)}",
        "Avg C-LAMAML SR%": round(float(np.mean([r["clamaml_sr"] for r in rows])) * 100, 2),
        "Avg Base 1 (Adapter) SR%": round(float(np.mean([r["b1_sr"] for r in rows])) * 100, 2),
        "Avg Base 2 (Concat) SR%": round(float(np.mean([r["b2_sr"] for r in rows])) * 100, 2),
    }

    xlsx_path = save_excel(args, rows, aggregate)
    
    print("\n" + "=" * 80)
    print("Overall comparison")
    for k, v in aggregate.items():
        print(f"  {k}: {v}")
    print(f"Results saved to: {xlsx_path}")
    print("=" * 80)

if __name__ == "__main__":
    main()
